#!/usr/bin/env python
# -*- coding:utf-8 -*-
'''
@time: 2016/12/1 14:11
@author: Silence
读取训练数据
'''
import openpyxl
from openpyxl import load_workbook
from openpyxl import worksheet
from openpyxl import Workbook
import numpy as np
def readExcel2Nparray(filename):
    wb = load_workbook(filename)#获取一个已经存在的excel文件wb
    sheetnames = wb.get_sheet_names()
    ws=wb.get_sheet_by_name(sheetnames[0])#打开该文件wb需要用到的worksheet即ws
    list1 = []
    for row in ws.iter_rows():
        b = list(a.value for a in row)
        list1.append(list(a.value for a in row))
        arr = np.array(list1)

    return arr

if __name__ == '__main__':
    print readExcel2Nparray(r'sum_negative_distance.xlsx')


